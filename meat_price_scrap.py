import requests
import pandas as pd
import numpy as np
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
import re as r

global dict1
global data_list

def extract_meat_part(prd_name):
    # [샘플]이 포함된 경우 None 반환
    if '[샘플]' in prd_name:
        return None
    # 접두어 제거 (예: [한정수량][업소용/슬라이스])
    cleaned_name = r.sub(r'^\[.*?\]', '', prd_name)

    # 부위 추출 (복합 부위명 포함)
    meat_part = r.findall(r'^([^-]+)', cleaned_name)
    meat_part_text = meat_part[0].split(']')[-1] if meat_part else '미상'
    
    if meat_part:
        # '/'가 포함된 경우 그대로 사용, 그렇지 않으면 첫 번째 매치만 사용
        return meat_part[0] if '/' in meat_part[0] else meat_part[0].split('/')[0]
    else:
        # 부위를 찾지 못한 경우 처리
        print(f"Warning: 부위 정보를 찾을 수 없음 - {prd_name}")
        return '미상'

def crawling(soup, part):
    items = soup.find_all("div", class_="detail_top")
    for item in items:
        try:
            prd_name = item.find('strong', class_='prd_name').text.replace(" ","")

            # [샘플] 제품 제외
            meat_part = extract_meat_part(prd_name)
            if meat_part is None:
                continue

            brand = r.findall('\\|([a-zA-Zㄱ-ㅎ가-힣]+)', prd_name)
            country = r.findall('-([a-zA-Zㄱ-ㅎ가-힣]+)(?=\\|)', prd_name)
            
            # 가격 정보 추출 수정
            price_info = item.find('em', class_='pric_stock')
            if price_info:
                price_text = price_info.text.replace(" ", "").replace(',', '')
                kg_price_match = r.search(r'\(kg당\)(\d+)', price_text)
                if kg_price_match:
                    price_text = kg_price_match.group(1)
                else:
                    g100_price_match = r.search(r'\(100g당\)(\d+)', price_text)
                    if g100_price_match:
                        price_text = str(int(g100_price_match.group(1)) * 10)
                    else:
                        price_match = r.search(r'(\d+)원', price_text)
                        if price_match:
                            price_text = price_match.group(1)
                        else:
                            price_text = ''
            else:
                price_text = ''

            grade = r.findall(r'-([A-Za-z0-9]+(?:그레이드)?)(?:\||$)', prd_name)
            if not grade:
                grade = r.findall(r'-((?:초이스|셀렉트|프라임|언그레이드|MB\d+(?:~\d+)?|MB\d+UP))(?:\||$)', prd_name)
            
            brand_text = brand[0] if brand else ''
            country_text = country[0] if country else ''
            grade_text = grade[-1] if grade else ''

            dict1 = {'기준일자': date, '조사업체': brand_text, '원산지': country_text, '부위': part, 
            '등급': grade_text, '가격KG': price_text}
            data_list.append(dict1)

            print(f"Name: {prd_name}")
            print(f"Brand: {brand_text}")
            print(f"Country: {country_text}")
            print(f"Grade: {grade_text}")
            print(f"Part: {part}")
            print(f"Price: {price_text}")
            print("---")

        except Exception as e:
            print(f"Error processing item: {e}")
            continue

    print(f"Processed items for {part}")

meat_part = {'목심','차돌양지','사태','앞다리%2F전각'}
file_path = '/Users/jaechankwon/Downloads/'
file_name = 'market_price.xlsx'
month = datetime.today().strftime("%m월")
date = datetime.today().strftime("%m-%d")
data_list = []

for part in meat_part:
    url = f"https://www.meatbox.co.kr/fo/product/productListPage.do?categorySeq=10005&articleCount=1000&itemCatName={part}&sellUnitType=B&listStyle=image"
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html5lib')
    crawling(soup, part)

# 엑셀 파일 열기 (기존 파일이 있으면 열고,  // 없으면 새로 만듦 // - 해당부분 주석처리 -)
# try:
workbook = openpyxl.load_workbook(file_path + file_name)
# except FileNotFoundError:
    # workbook = openpyxl.Workbook()

# 현재 월의 시트가 있는지 확인
if month in workbook.sheetnames:
    sheet = workbook[month]
    existing_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_data.append(dict(zip(['기준일자', '조사업체', '원산지', '부위', '등급', '가격KG'], row)))
    
    # 기존 데이터와 새 데이터를 합칩니다
    temp_data = existing_data + data_list
else:
    # 새 시트를 가장 오른쪽에 추가
    sheet = workbook.create_sheet(title=month)
    workbook.active = sheet
    temp_data = data_list

# 데이터프레임 생성
df = pd.DataFrame(temp_data)

# 기존 데이터 삭제
for row in sheet['A2:F' + str(sheet.max_row)]:
    for cell in row:
        cell.value = None

# 데이터를 시트에 쓰기 (기존 데이터 + 새 데이터)
for r, row in enumerate(df.values, start=2):
    for c, value in enumerate(row, start=1):
        sheet.cell(row=r, column=c, value=value)

# 열 이름 쓰기
for c, column_name in enumerate(df.columns, start=1):
    sheet.cell(row=1, column=c, value=column_name)

# 파일 저장
workbook.save(file_path + file_name)

print("저장완료")