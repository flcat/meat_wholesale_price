import asyncio
from playwright.async_api import async_playwright
import aiohttp
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import openpyxl
import yaml
import logging
import re

# 로깅 설정
logging.basicConfig(filename='meat_price_scrap.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 설정 파일 로드
with open('/Users/jaechankwon/meat_market_price/config2.yaml', 'r') as f:
    config = yaml.safe_load(f)

def extract_meat_part(prd_name):
    for part in config['meat_parts']:
        if part['name'] in prd_name:
            return part['name']
    return '기타'

async def retry_goto(page, url, max_attempts=3):
    for attempt in range(max_attempts):
        try:
            await page.goto(url, wait_until="load", timeout=60000)
            return
        except TimeoutError:
            if attempt == max_attempts - 1:
                raise
            print(f"Timeout on attempt {attempt + 1}, retrying...")
            await asyncio.sleep(5) 

async def crawl_category(page, category, meat_part):
    url = f"{config['base_url']}{config['product_list_endpoint']}"
    params = {
        'dispCtgNo': category['params']['dispCtgNo'],
        'leafCtgNo': meat_part['leafCtgNo'],
        'dispCtgNoList': meat_part['dispCtgNoList']
    }
    query_string = '&'.join([f"{k}={v}" for k, v in params.items()])
    full_url = f"{url}?{query_string}"
    
    try:
        await page.goto(full_url, wait_until="load", timeout=60000)
        await retry_goto(page, full_url)
        # 페이지 로드 상태 확인 및 스크롤
        await page.evaluate("""
            () => {
                return new Promise((resolve) => {
                    let totalHeight = 0;
                    let distance = 100;
                    let timer = setInterval(() => {
                        let scrollHeight = document.body.scrollHeight;
                        window.scrollBy(0, distance);
                        totalHeight += distance;
                        if(totalHeight >= scrollHeight){
                            clearInterval(timer);
                            resolve();
                        }
                    }, 100);
                });
            }
        """)
        
        await page.wait_for_selector("#content > div.inner.mo-wide > div.list-wrap > div.list-area > div.list-cont > ul > li", timeout=30000)
        
        content = await page.content()
        soup = BeautifulSoup(content, 'html.parser')
        items = soup.select("#content > div.inner.mo-wide > div.list-wrap > div.list-area > div.list-cont > ul > li")
        
        if not items:
            print(f"No items found for {category['name']} - {meat_part['name']}")
            return []

        category_data = []
        print(f"Crawling {category['name']} - {meat_part['name']}...")
        
        for item in items:
            try:
                prd_info = item.select_one("div > a > div.pr-info > div.name-wrap > p")
                prd_name = prd_info.text.strip() if prd_info else ""
                
                # 정규표현식을 사용하여 브랜드와 부위 분리
                match = re.match(r'(.*?)_(.+)', prd_name)
                if match:
                    brand, part = match.groups()
                else:
                    brand, part = "Unknown", prd_name

                grade_elem = item.select_one("div > div > div:nth-child(2) > div > p")
                grade = grade_elem.text.strip() if grade_elem else ""
                
                price_elem = item.select_one("div > div > div.pd-row.md.show-pc > p > span.pd-price.xs.c-primary")
                price = price_elem.text.strip().replace(',', '').replace('원', '') if price_elem else ""
                
                dict1 = {
                    '기준일자': datetime.today().strftime("%m-%d"),
                    '조사업체': brand,
                    '원산지': category['name'],
                    '부위': meat_part['name'],
                    '등급': grade,
                    '가격KG': price
                }
                category_data.append(dict1)
                
                print(f"Processed: {prd_name}")
                
            except Exception as e:
                logging.error(f"Error processing item: {e}")
                print(f"Error processing item: {e}")

        print(f"Completed crawling {category['name']} - {meat_part['name']}. Total items: {len(category_data)}")
        logging.info(f"Processed {len(category_data)} items for {category['name']} - {meat_part['name']}")
        return category_data

    except Exception as e:
        logging.error(f"Error crawling {category['name']} - {meat_part['name']}: {e}")
        print(f"Error crawling {category['name']} - {meat_part['name']}: {e}")
        return []

async def main():
    try:
        logging.info("Starting the main process")
        all_data = []
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            )
            
            tasks = []
            for category in config['categories']:
                for meat_part in config['meat_parts']:
                    page = await context.new_page()
                    task = asyncio.create_task(crawl_category(page, category, meat_part))
                    tasks.append(task)
            
            results = await asyncio.gather(*tasks)
            for result in results:
                all_data.extend(result)
            
            await browser.close()

        logging.info(f"Crawling process completed. Total items collected: {len(all_data)}")
        print(f"Crawling process completed. Total items collected: {len(all_data)}")

        # 엑셀 파일 처리 부분
        month = datetime.today().strftime("%m월")
        try:
            workbook = openpyxl.load_workbook(config['file_path'] + config['file_name'])
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        if month in workbook.sheetnames:
            sheet = workbook[month]
            existing_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                existing_data.append(dict(zip(['기준일자', '조사업체', '원산지', '부위', '등급', '가격KG'], row)))
            
            # 중복 제거를 위해 기존 데이터와 새 데이터를 합치고 중복 제거
            all_data = existing_data + all_data
            df = pd.DataFrame(all_data)
            df = df.drop_duplicates(subset=['기준일자', '조사업체', '원산지', '부위', '등급'], keep='last')
            temp_data = df.to_dict('records')
        else:
            sheet = workbook.create_sheet(title=month)
            workbook.active = sheet
            temp_data = all_data

        # 기존 데이터 삭제
        for row in sheet['A2:F' + str(sheet.max_row)]:
            for cell in row:
                cell.value = None

        # 새 데이터 쓰기
        for r, row in enumerate(temp_data, start=2):
            for c, (column, value) in enumerate(row.items(), start=1):
                sheet.cell(row=r, column=c, value=value)

        # 열 이름 쓰기
        columns = ['기준일자', '조사업체', '원산지', '부위', '등급', '가격KG']
        for c, column_name in enumerate(columns, start=1):
            sheet.cell(row=1, column=c, value=column_name)

        workbook.save(config['file_path'] + config['file_name'])
        logging.info("Data saved successfully")
        print("Data saved successfully. Process complete.")

    except Exception as e:
        logging.error(f"Error in main process: {e}")
        print(f"Error in main process: {e}")

if __name__ == "__main__":
    asyncio.run(main())