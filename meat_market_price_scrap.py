import requests
import pandas as pd
import numpy as np
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
import re as r
import yaml
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

# 로깅 설정
logging.basicConfig(filename='meat_price_scrap.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 설정 파일 로드
with open('/Users/jaechankwon/meat_market_price/config.yaml', 'r') as f:
    config = yaml.safe_load(f)

def extract_meat_part(prd_name):
    # [샘플]이 포함된 경우 None 반환
    if '[샘플]' in prd_name:
        return None
    # 접두어 제거 (예: [한정수량][업소용/슬라이스])
    cleaned_name = r.sub(r'^\[.*?\]', '', prd_name)

    # 부위 추출 (복합 부위명 포함)
    meat_part = r.findall(r'^([^-]+)', cleaned_name)
    meat_part_text = meat_part[0].split(']')[-1] if meat_part else '미상'
    
    if meat_part:
        # '/'가 포함된 경우 그대로 사용, 그렇지 않으면 첫 번째 매치만 사용
        return meat_part[0] if '/' in meat_part[0] else meat_part[0].split('/')[0]
    else:
        # 부위를 찾지 못한 경우 처리
        logging.warning(f"Warning: 부위 정보를 찾을 수 없음 - {prd_name}")
        print(f"Warning: 부위 정보를 찾을 수 없음 - {prd_name}")
        return '미상'

def crawling_with_retry(part, max_retries=3, delay=5):
    url = f"{config['url_base']}?categorySeq=10005&articleCount=1000&itemCatName={part}&sellUnitType=B&listStyle=image"
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html5lib')
            items = soup.find_all("div", class_="detail_top")
            part_data = []

            print(f"Crawling {part}...")

            for item in items:
                try:
                    prd_name = item.find('strong', class_='prd_name').text.replace(" ","")

                    # [샘플] 제품 제외
                    meat_part = extract_meat_part(prd_name)
                    if meat_part is None:
                        continue

                    brand = r.findall('\\|([a-zA-Zㄱ-ㅎ가-힣]+)', prd_name)
                    country = r.findall('-([a-zA-Zㄱ-ㅎ가-힣]+)(?=\\|)', prd_name)
                    
                    # 가격 정보 추출 수정
                    price_info = item.find('em', class_='pric_stock')
                    if price_info:
                        price_text = price_info.text.replace(" ", "").replace(',', '')
                        kg_price_match = r.search(r'\(kg당\)(\d+)', price_text)
                        if kg_price_match:
                            price_text = kg_price_match.group(1)
                        else:
                            g100_price_match = r.search(r'\(100g당\)(\d+)', price_text)
                            if g100_price_match:
                                price_text = str(int(g100_price_match.group(1)) * 10)
                            else:
                                price_match = r.search(r'(\d+)원', price_text)
                                if price_match:
                                    price_text = price_match.group(1)
                                else:
                                    price_text = ''
                    else:
                        price_text = ''

                    grade = r.findall(r'-([A-Za-z0-9]+(?:그레이드)?)(?:\||$)', prd_name)
                    if not grade:
                        grade = r.findall(r'-((?:초이스|셀렉트|프라임|언그레이드|MB\d+(?:~\d+)?|MB\d+UP))(?:\||$)', prd_name)
                    
                    brand_text = brand[0] if brand else ''
                    country_text = country[0] if country else ''
                    grade_text = grade[-1] if grade else ''

                    dict1 = {'기준일자': datetime.today().strftime("%m-%d"), '조사업체': brand_text, '원산지': country_text, '부위': part, 
                    '등급': grade_text, '가격KG': price_text}
                    part_data.append(dict1)

                    print(f"Processed: {prd_name}")

                except Exception as e:
                    logging.error(f"Error processing item for {part}: {e}")
                    print(f"Error processing item: {e}")

            print(f"Completed crawling {part}. Total items: {len(part_data)}")
            logging.info(f"Processed {len(part_data)} items for {part}")
            return part_data

        except requests.RequestException as e:
            logging.warning(f"Attempt {attempt + 1} failed for {part}: {e}")
            print(f"Attempt {attempt + 1} failed for {part}: {e}")
            if attempt < max_retries - 1:
                time.sleep(delay)
            else:
                logging.error(f"All attempts failed for {part}")
                print(f"All attempts failed for {part}")
                return []

def main():
    month = datetime.today().strftime("%m월")
    date = datetime.today().strftime("%m-%d")
    data_list = []

    print("Starting crawling process...")

    # 병렬 처리로 크롤링 수행
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_part = {executor.submit(crawling_with_retry, part): part for part in config['meat_parts']}
        for future in as_completed(future_to_part):
            part = future_to_part[future]
            try:
                part_data = future.result()
                data_list.extend(part_data)
                print(f"Completed processing {part}")
            except Exception as e:
                logging.error(f"Error processing {part}: {e}")
                print(f"Error processing {part}: {e}")

    print("Crawling process completed. Saving data to Excel...")

    # 엑셀 파일 처리
    try:
        workbook = openpyxl.load_workbook(config['file_path'] + config['file_name'])
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    if month in workbook.sheetnames:
        sheet = workbook[month]
        existing_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            existing_data.append(dict(zip(['기준일자', '조사업체', '원산지', '부위', '등급', '가격KG'], row)))
        temp_data = existing_data + data_list
    else:
        sheet = workbook.create_sheet(title=month)
        workbook.active = sheet
        temp_data = data_list

    df = pd.DataFrame(temp_data)

    for row in sheet['A2:F' + str(sheet.max_row)]:
        for cell in row:
            cell.value = None

    for r, row in enumerate(df.values, start=2):
        for c, value in enumerate(row, start=1):
            sheet.cell(row=r, column=c, value=value)

    for c, column_name in enumerate(df.columns, start=1):
        sheet.cell(row=1, column=c, value=column_name)

    workbook.save(config['file_path'] + config['file_name'])
    logging.info("Data saved successfully")
    print("Data saved successfully. Process complete.")

if __name__ == "__main__":
    main()